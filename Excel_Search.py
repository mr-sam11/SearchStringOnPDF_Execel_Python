from openpyxl import load_workbook


def search_excel(excel_file, search_string):
    # Load the Excel workbook
    workbook = load_workbook(excel_file)
    # This Loop is for multiple sheets on excel
    for sheet in workbook:
        # This Loop is for current sheet and iterate through the rows in the sheet
        for row_num, row in enumerate(sheet.iter_rows()):
            # this Loop is for the cells in the row
            for col_num, cell in enumerate(row):
                # This conditional statement for if string is found
                if cell.value == search_string:
                    print(f"String is '{search_string}' found in row {row_num + 1} column {col_num + 1}")


# Program For Search string on Excel and display the row and col

excel_file = 'excelsample.xlsx'
search_string = input("Enter String Which You Want To Search  = ");
search_excel(excel_file, search_string)
